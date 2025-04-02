from django.shortcuts import render, redirect, get_object_or_404
from .models import Student, Subject
from django.http import HttpResponse
from django.core.paginator import Paginator
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.utils import timezone
from django.core.exceptions import ValidationError


def index(request):
    return render(request, 'index.html')
# --- Student Views ---
#  Detail for students 
# Show student detail
def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    subjects = Subject.objects.filter(student=student) .order_by('year')  # Fetch subjects for this student
    
    total_marks = 0
    for subject in subjects:
        # Sum marks from each subject
        total_marks += (subject.subject1_marks + subject.subject2_marks +
                        subject.subject3_marks + subject.subject4_marks +
                        subject.subject5_marks + subject.subject6_marks)
    
    return render(request, 'student/student_detail.html', {'student': student, 'subjects': subjects, 'total_marks': total_marks})

# List all students
def student_list(request):
    query = request.GET.get('q', '')  # Get the search query from the URL
    students = Student.objects.all()

    if query:
        students = students.filter(name__icontains=query)  # Filter students by name (case-insensitive)

    paginator = Paginator(students, 3)  # Show 10 students per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'student/student_list.html', {'page_obj': page_obj, 'query': query})

# Create a new student
def create_student(request):
    if request.method == 'POST':
        # Get data from form
        name = request.POST.get('name')
        dob = request.POST.get('dob')
        nrc = request.POST.get('nrc')
        father_name = request.POST.get('father_name')
        mother_name = request.POST.get('mother_name')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        grade = request.POST.get('grade')
        country = request.POST.get('country')

        nrc_exists = Student.objects.filter(nrc=nrc).exists()
        if nrc_exists:
            messages.error(request, "A student with this NRC number already exists.")
            return redirect('create_student')  # Redirect back to the form

        # Validation (basic)
        if not name or not dob or not nrc or not father_name or not mother_name or not phone or not address or not grade or not country:
            messages.error(request, "Please fill in all the required fields.")
            return redirect('create_student')  # Redirect back to form

        # Save Student
        student = Student(
            name=name,
            dob=dob,
            nrc=nrc,
            father_name=father_name,
            mother_name=mother_name,
            phone=phone,
            address=address,
            grade=grade,
            country=country
        )
        student.save()

        # Handle multiple subject entries
        years = request.POST.getlist('new_subject_year')
        subject1_marks = request.POST.getlist('new_subject1_marks')
        subject2_marks = request.POST.getlist('new_subject2_marks')
        subject3_marks = request.POST.getlist('new_subject3_marks')
        subject4_marks = request.POST.getlist('new_subject4_marks')
        subject5_marks = request.POST.getlist('new_subject5_marks')
        subject6_marks = request.POST.getlist('new_subject6_marks')

        # Iterate through the submitted subject data and create Subject objects
        for i in range(len(years)):
            year = years[i]
            marks1 = subject1_marks[i]
            marks2 = subject2_marks[i]
            marks3 = subject3_marks[i]
            marks4 = subject4_marks[i]
            marks5 = subject5_marks[i]
            marks6 = subject6_marks[i]

            # You might want to add validation here to ensure marks are not empty
            if year and marks1 and marks2 and marks3 and marks4 and marks5 and marks6:
                Subject.objects.create(
                    student=student,
                    year=year,
                    subject1_marks=marks1,
                    subject2_marks=marks2,
                    subject3_marks=marks3,
                    subject4_marks=marks4,
                    subject5_marks=marks5,
                    subject6_marks=marks6
                )

        # Display success message
        messages.success(request, "Student successfully created!")
        return redirect('student_list')  # Optionally, redirect to another page after success

    return render(request, 'student/create_student.html') # Render a form for student creation

def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    existing_subjects = Subject.objects.filter(student=student)  # Get all subjects for the student

    if request.method == "POST":
        try:
            # Handle the main student data update
            student.name = request.POST.get('name')
            student.dob = request.POST.get('dob')
            student.nrc = request.POST.get('nrc')
            student.father_name = request.POST.get('father_name')
            student.mother_name = request.POST.get('mother_name')
            student.phone = request.POST.get('phone')
            student.address = request.POST.get('address')
            student.grade = request.POST.get('grade')
            student.country = request.POST.get('country')

            # Now handle the subject marks update for each existing subject
            for subject in existing_subjects:
                subject.subject1_marks = float(request.POST.get(f'subject1_marks_{subject.year}'))
                subject.subject2_marks = float(request.POST.get(f'subject2_marks_{subject.year}'))
                subject.subject3_marks = float(request.POST.get(f'subject3_marks_{subject.year}'))
                subject.subject4_marks = float(request.POST.get(f'subject4_marks_{subject.year}'))
                subject.subject5_marks = float(request.POST.get(f'subject5_marks_{subject.year}'))
                subject.subject6_marks = float(request.POST.get(f'subject6_marks_{subject.year}'))
                subject.save()  # Save the updated subject marks

            # Handle multiple new subject entries
            years = request.POST.getlist('new_subject_year')
            subject1_marks = request.POST.getlist('new_subject1_marks')
            subject2_marks = request.POST.getlist('new_subject2_marks')
            subject3_marks = request.POST.getlist('new_subject3_marks')
            subject4_marks = request.POST.getlist('new_subject4_marks')
            subject5_marks = request.POST.getlist('new_subject5_marks')
            subject6_marks = request.POST.getlist('new_subject6_marks')

            # Iterate through the submitted subject data and create new Subject objects
            for i in range(len(years)):
                year = years[i]
                marks1 = subject1_marks[i]
                marks2 = subject2_marks[i]
                marks3 = subject3_marks[i]
                marks4 = subject4_marks[i]
                marks5 = subject5_marks[i]
                marks6 = subject6_marks[i]

                # Ensure all data is present and valid
                if year and marks1 and marks2 and marks3 and marks4 and marks5 and marks6:
                    try:
                        Subject.objects.create(
                            student=student,
                            year=year,
                            subject1_marks=float(marks1),
                            subject2_marks=float(marks2),
                            subject3_marks=float(marks3),
                            subject4_marks=float(marks4),
                            subject5_marks=float(marks5),
                            subject6_marks=float(marks6)
                        )
                    except ValueError:
                        raise ValidationError("Invalid subject marks provided.")

            student.save()  # Save the updated student data

            # Redirect to the student list page or profile page
            return redirect('student_list')  # Adjust this URL if necessary

        except (ValueError, ValidationError) as e:
            error_message = "Invalid data: " + str(e)
            return render(request, 'student/edit_student.html', {
                'student': student, 
                'existing_subjects': existing_subjects, 
                'error_message': error_message
            })

    return render(request, 'student/edit_student.html', {
        'student': student, 
        'existing_subjects': existing_subjects
    })
# Delete a student
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    student.delete()
    return redirect('student_list')  # After deletion, redirect to student list


# --- Subject Views ---

# List all subjects
def subject_list(request):
    query = request.GET.get('q', '')  # Get the search query from the URL
    subjects = Subject.objects.all().order_by('year')  # Order by year in ascending order

    

    if query:
        subjects = subjects.filter(student__name__icontains=query)  # Filter subjects by student name

    paginator = Paginator(subjects, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'subject/subject_list.html', {'page_obj': page_obj, 'query': query})

# Create a new subject
def create_subject(request):
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        year = request.POST.get('year')
        subject1_marks = request.POST.get('subject1_marks')
        subject2_marks = request.POST.get('subject2_marks')
        subject3_marks = request.POST.get('subject3_marks')
        subject4_marks = request.POST.get('subject4_marks')
        subject5_marks = request.POST.get('subject5_marks')
        subject6_marks = request.POST.get('subject6_marks')

        student = get_object_or_404(Student, id=student_id)
        subject = Subject(student=student, year=year, subject1_marks=subject1_marks,
                          subject2_marks=subject2_marks, subject3_marks=subject3_marks,
                          subject4_marks=subject4_marks, subject5_marks=subject5_marks,
                          subject6_marks=subject6_marks)
        subject.save()
        return redirect('subject_list')  # After creation, redirect to subject list
    
    students = Student.objects.all()  # To show students in the create subject form
    return render(request, 'subject/create_subject.html', {'students': students})

# Edit an existing subject
def edit_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    
    if request.method == "POST":
        subject.year = request.POST.get('year')
        subject.subject1_marks = request.POST.get('subject1_marks')
        subject.subject2_marks = request.POST.get('subject2_marks')
        subject.subject3_marks = request.POST.get('subject3_marks')
        subject.subject4_marks = request.POST.get('subject4_marks')
        subject.subject5_marks = request.POST.get('subject5_marks')
        subject.subject6_marks = request.POST.get('subject6_marks')
        subject.save()
        return redirect('subject_list')  # After edit, redirect to subject list
    
    return render(request, 'subject/edit_subject.html', {'subject': subject})  # Show subject data in form

# Delete a subject
def delete_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    subject.delete()
    return redirect('subject_list')  # After deletion, redirect to subject list

# --- Export Students to XLSX ---
def export_students(request):
    # Get all students
    students = Student.objects.all()

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Define the headers
    headers = ['ID', 'Name', 'DOB', 'NRC', 'Father Name', 'Mother Name', 'Phone', 'Address', 'Grade', 'Country']
    ws.append(headers)

    # Add student data to the Excel sheet
    for student in students:
        ws.append([student.id, student.name, student.dob, student.nrc, student.father_name,
                   student.mother_name, student.phone, student.address, student.grade, student.country])

    # Set the column widths
    for col in range(1, len(headers) + 1):
        column = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Return the file as an HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    wb.save(response)
    return response


# --- Import Students from XLSX ---
def import_students(request):
    if request.method == "POST" and request.FILES["excel_file"]:
        excel_file = request.FILES["excel_file"]
        fs = FileSystemStorage()
        file_path = fs.save(excel_file.name, excel_file)
        file_url = fs.url(file_path)

        # Open the uploaded Excel file
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Iterate through the rows and insert the data into the database
        for row in ws.iter_rows(min_row=2, values_only=True):
            student_data = {
                'id': row[0],
                'name': row[1],
                'dob': row[2],
                'nrc': row[3],
                'father_name': row[4],
                'mother_name': row[5],
                'phone': row[6],
                'address': row[7],
                'grade': row[8],
                'country': row[9],
            }
            # Create or update the student data in the database
            Student.objects.update_or_create(id=student_data['id'], defaults=student_data)

        # Provide a success message
        messages.success(request, "Students imported successfully!")
        return redirect('student_list')

    return render(request, 'student/import_students.html')

# Import subjects from XLSX
def import_subjects(request):
    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            student_name, year, subject, subject1_marks, subject2_marks, subject3_marks, subject4_marks, subject5_marks, subject6_marks, total_marks = row
            
            # Fetch the student based on the name (assuming the name is unique, else modify to use another identifier)
            try:
                student = Student.objects.get(name=student_name)
            except Student.DoesNotExist:
                messages.error(request, f"Student {student_name} not found in the database.")
                continue  # Skip the current row if the student does not exist

            # Create a new Subject record
            Subject.objects.create(
                student=student,
                year=year,
                subject1_marks=subject1_marks,
                subject2_marks=subject2_marks,
                subject3_marks=subject3_marks,
                subject4_marks=subject4_marks,
                subject5_marks=subject5_marks,
                subject6_marks=subject6_marks
            )
        
        messages.success(request, 'Subjects imported successfully!')
    
    return render(request, 'subject/import_subject.html')


# Export Subjects to XLSX
def export_subjects(request):
    subjects = Subject.objects.select_related('student')  # Fetch related student data efficiently

    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subjects"

    # Add column headers
    headers = ["Student Name", "Year", "Subject", "Marks 1", "Marks 2", "Marks 3", "Marks 4", "Marks 5", "Marks 6", "Total Marks"]
    for col_num, header in enumerate(headers, 1):
        ws[get_column_letter(col_num) + '1'] = header

    # Add subject data rows
    for row_num, subject in enumerate(subjects, 2):
        ws[get_column_letter(1) + str(row_num)] = subject.student.name  # Add student name
        ws[get_column_letter(2) + str(row_num)] = subject.year
        ws[get_column_letter(3) + str(row_num)] = f"Subject {row_num - 1}"
        ws[get_column_letter(4) + str(row_num)] = subject.subject1_marks
        ws[get_column_letter(5) + str(row_num)] = subject.subject2_marks
        ws[get_column_letter(6) + str(row_num)] = subject.subject3_marks
        ws[get_column_letter(7) + str(row_num)] = subject.subject4_marks
        ws[get_column_letter(8) + str(row_num)] = subject.subject5_marks
        ws[get_column_letter(9) + str(row_num)] = subject.subject6_marks
        ws[get_column_letter(10) + str(row_num)] = subject.total_marks

    # Response with the XLSX file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=subjects.xlsx"
    wb.save(response)
    return response